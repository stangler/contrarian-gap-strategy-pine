import csv
from pathlib import Path
from openpyxl import load_workbook

def extract_from_csv(csv_file):
    data = {}
    with open(csv_file, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        first_row = next(reader, None)
        if first_row and len(first_row) >= 1:
            data['銘柄'] = first_row[0].strip()
        next(reader, None)  # ヘッダー行スキップ
        for row in reader:
            if len(row) >= 2:
                key = row[0].strip()
                value = row[1].strip()
                if key not in data:
                    data[key] = value
    return data

def update_excel(csv_data_list):
    wb = load_workbook("format.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        symbol_cell = row[1]  # 列B
        if not symbol_cell.value:
            continue

        symbol = str(symbol_cell.value).strip()
        csv_data = next((d for d in csv_data_list if d.get('銘柄') == symbol), None)

        if csv_data:
            # 既存列（C〜I = index 2〜8）
            for col, key in zip(range(2, 9), [
                '総損益', '最大ドローダウン', 'トレード総数',
                '勝ちトレード', '負けトレード', '勝率', 'プロフィットファクター'
            ]):
                row[col].value = csv_data.get(key)

            # GU/GD/Cont列（J〜R = index 9〜17）
            for col, key in zip(range(9, 18), [
                'GU_勝', 'GU_負', 'GU_勝率',
                'GD_勝', 'GD_負', 'GD_勝率',
                'Cont_勝', 'Cont_負', 'Cont_勝率',
            ]):
                row[col].value = csv_data.get(key)

            print(f"  ✓ {symbol}")
        else:
            print(f"  ✗ {symbol}: CSVなし")

    wb.save("format.xlsx")
    print("\nSaved to format.xlsx")

def main():
    csv_files = list(Path(".").glob("*.csv"))
    print(f"Found {len(csv_files)} CSV files")

    csv_data_list = [extract_from_csv(f) for f in csv_files]
    update_excel(csv_data_list)

if __name__ == "__main__":
    main()
