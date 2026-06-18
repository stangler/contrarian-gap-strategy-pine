import csv
from pathlib import Path
from openpyxl import load_workbook

def extract_from_csv(csv_file):
    data = {'_file': str(csv_file)}
    with open(csv_file, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        first_row = next(reader, None)
        if first_row and len(first_row) >= 1:
            data['銘柄'] = first_row[0].strip()
        next(reader, None)  # ヘッダー行スキップ
        for row in reader:
            if len(row) >= 2:
                key = row[0].strip()
                value = row[1].strip()
                if key not in data:
                    data[key] = value
    return data

def slot_key_from_data(data: dict) -> str:
    """CSV内のSlot開始(分)/Slot終了(分)からスロット識別子を生成。
    複数スロットのCSVが同じフォルダに混在していても、これでグループ分けできる。"""
    try:
        start = int(float(data.get('Slot開始', '')))
        end = int(float(data.get('Slot終了', '')))
    except (TypeError, ValueError):
        return "unknown"
    def hhmm(m):
        h, mi = divmod(m, 60)
        return f"{h:02d}{mi:02d}"
    return f"{hhmm(start)}_{hhmm(end)}"

def update_excel(csv_data_list, template="format.xlsx", output_name="format.xlsx"):
    wb = load_workbook(template)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        symbol_cell = row[1]  # 列B
        if not symbol_cell.value:
            continue

        symbol = str(symbol_cell.value).strip()
        csv_data = next((d for d in csv_data_list if d.get('銘柄') == symbol), None)

        if csv_data:
            # 既存列（C〜I = index 2〜8）
            for col, key in zip(range(2, 9), [
                '総損益', '最大ドローダウン', 'トレード総数',
                '勝ちトレード', '負けトレード', '勝率', 'プロフィットファクター'
            ]):
                row[col].value = csv_data.get(key)

            # GU/GD/Cont列（J〜R = index 9〜17）
            for col, key in zip(range(9, 18), [
                'GU_勝', 'GU_負', 'GU_勝率',
                'GD_勝', 'GD_負', 'GD_勝率',
                'Cont_勝', 'Cont_負', 'Cont_勝率',
            ]):
                row[col].value = csv_data.get(key)

            print(f"  ✓ {symbol}")
        else:
            print(f"  ✗ {symbol}: CSVなし")

    wb.save(output_name)
    print(f"  Saved to {output_name}")

def main():
    csv_files = sorted(Path(".").glob("*.csv"), key=lambda p: p.stat().st_mtime)
    print(f"Found {len(csv_files)} CSV files")

    csv_data_list = [extract_from_csv(f) for f in csv_files]

    # スロット別にグループ分け（同フォルダに複数スロットのCSVが混在していてもOK）
    # 同一銘柄+スロットのCSVが複数（別日付など）ある場合は、mtimeが新しい方を採用
    # （csv_filesをmtime昇順にしてあるので、辞書への代入で後勝ち＝最新が残る）
    grouped: dict[str, dict] = {}
    for d in csv_data_list:
        slot = slot_key_from_data(d)
        symbol = d.get('銘柄')
        bucket = grouped.setdefault(slot, {})
        if symbol in bucket:
            print(f"  ⚠ 重複検出: {symbol}（スロット{slot}） "
                  f"{bucket[symbol]['_file']} → {d['_file']} を採用（新しい方）")
        bucket[symbol] = d

    print(f"検出スロット: {list(grouped.keys())}\n")

    for slot, symbol_map in grouped.items():
        items = list(symbol_map.values())
        if slot == "unknown":
            out_name = "format.xlsx"  # スロット情報なしCSV（旧形式）は従来通り
        else:
            out_name = f"format_{slot}.xlsx"
        print(f"--- スロット「{slot}」（{len(items)}件） → {out_name} ---")
        update_excel(items, template="format.xlsx", output_name=out_name)
        print()

if __name__ == "__main__":
    main()
