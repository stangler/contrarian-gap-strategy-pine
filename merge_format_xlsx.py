"""
複数スロットの format_HHMM_HHMM.xlsx を1ファイルに統合する。
extract_data_contrarian.py 実行後、同フォルダで実行すること。

入力: format_0900_0930.xlsx, format_0930_1000.xlsx, ... (同フォルダ内すべて)
出力: format_all_slots.xlsx（1シートに全スロットを縦に結合、「スロット」列を追加）
"""
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook

PATTERN = re.compile(r"^format_(\d{4})_(\d{4})\.xlsx$")
OUTPUT_FILE = "format_all_slots.xlsx"

def hhmm_to_clock(s: str) -> str:
    return f"{s[:2]}:{s[2:]}"

def main():
    files = sorted(p for p in Path(".").glob("format_*.xlsx") if PATTERN.match(p.name))
    print(f"対象ファイル {len(files)}件: {[f.name for f in files]}\n")

    if not files:
        print("✗ format_HHMM_HHMM.xlsx 形式のファイルが見つからない（extract_data_contrarian.pyを先に実行）")
        return

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "統合"

    header_written = False
    total_rows = 0

    for f in files:
        m = PATTERN.match(f.name)
        slot_label = f"{hhmm_to_clock(m.group(1))}-{hhmm_to_clock(m.group(2))}"

        wb = load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"  ⚠ {f.name}: 空シート、スキップ")
            continue

        if not header_written:
            out_ws.append(list(rows[0]) + ["スロット"])
            header_written = True

        count = 0
        for row in rows[1:]:
            out_ws.append(list(row) + [slot_label])
            count += 1

        print(f"  ✓ {f.name} → スロット「{slot_label}」 {count}行")
        total_rows += count

    out_wb.save(OUTPUT_FILE)
    print(f"\n合計 {total_rows}行 → {OUTPUT_FILE} に保存")

if __name__ == "__main__":
    main()
