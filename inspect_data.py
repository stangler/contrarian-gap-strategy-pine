import sys
import os
from pathlib import Path

# CSVファイルの確認
print("=== CSV Files ===")
csv_files = list(Path(".").glob("*.csv"))
if csv_files:
    for f in csv_files[:3]:  # 最初の3つだけ表示
        print(f"\nFile: {f}")
        try:
            with open(f, encoding="utf-8-sig") as fp:
                lines = fp.readlines()
                for i, line in enumerate(lines[:10]):
                    print(f"  {i+1}: {line.strip()}")
        except Exception as e:
            print(f"  Error reading: {e}")
else:
    print("No CSV files found")

# format.xlsxの確認
print("\n=== format.xlsx ===")
try:
    import openpyxl
    wb = openpyxl.load_workbook("format.xlsx")
    print(f"Sheets: {wb.sheetnames}")
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    
    # ヘッダー行
    print("\nHeader row (row 1):")
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
        print(f"  Column {cell.column}: {cell.value}")
    
    # 最初の数行
    print("\nFirst 3 data rows:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4), start=2):
        values = [cell.value for cell in row]
        print(f"  Row {row_idx}: {values}")
        
except ImportError:
    print("openpyxl not installed. Please install with: pip install openpyxl")
except FileNotFoundError:
    print("format.xlsx not found")
except Exception as e:
    print(f"Error: {e}")