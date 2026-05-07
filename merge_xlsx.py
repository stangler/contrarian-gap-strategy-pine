"""Merge all xlsx files in the xlsx folder into a single xlsx file."""

import os
import openpyxl
from openpyxl import Workbook

XLSX_DIR = "xlsx"
OUTPUT_FILE = "merged.xlsx"

# 前場・後場の順に並べる
file_order = [
    "0900-0930.xlsx",
    "0930-1000.xlsx",
    "1000-1030.xlsx",
    "1030-1100.xlsx",
    "1100-1130.xlsx",
    "1230-1300.xlsx",
    "1300-1330.xlsx",
    "1330-1400.xlsx",
    "1400-1430.xlsx",
    "1430-1500.xlsx",
    "1500-1525.xlsx",
]

files = []
for fname in file_order:
    fpath = os.path.join(XLSX_DIR, fname)
    if os.path.exists(fpath):
        files.append(fname)
    else:
        print(f"Warning: {fname} not found, skipping.")

if not files:
    print("No xlsx files found.")
    exit(1)

# Read all data
all_rows = []
headers = None

for fname in files:
    fpath = os.path.join(XLSX_DIR, fname)
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if headers is None:
        headers = list(rows[0]) + ["時間帯"]  # Add time period column
        all_rows.append(headers)

    for row in rows[1:]:
        all_rows.append(list(row) + [fname.replace(".xlsx", "")])

# Write merged workbook
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "merged"

for row_idx, row_data in enumerate(all_rows, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_out.cell(row=row_idx, column=col_idx, value=value)

wb_out.save(OUTPUT_FILE)
print(f"Merged {len(files)} files into {OUTPUT_FILE}")
print(f"Total rows (including header): {len(all_rows)}")